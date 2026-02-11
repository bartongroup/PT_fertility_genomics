#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make_master_fertility_workbook_from_results.py

Create a single Excel workbook that consolidates fertility-relevant evidence
from the project's results folder structure.

This script assumes the following results layout:

results/
  01_hpo/
    male_infertility_genes_summary.tsv
    gene_lists/male_infertility_gene_symbols.tsv
  04_clinvar_collapsed/
    male_infertility_clinvar_variants_best.tsv
    male_infertility_clinvar_variants_best_pathogenic.tsv
  05_clinvar_high_confidence/
    male_infertility_clinvar_variants_high_confidence.tsv
    male_infertility_clinvar_variants_high_confidence_pathogenic.tsv
  06_reports/
    male_infertility_clinvar_pathogenic_high_confidence_gene_summary.tsv
    male_infertility_clinvar_pathogenic_high_confidence_report.tsv
  <testis_run_id>/
    01_gtex_specificity/gtex_v11_tissue_medians.tsv
    08_proteomics_annotation/gtex_v11_testis_ranked_with_sperm_presence_function_hpo_prot.tsv
    09_high_confidence_final/high_confidence_testis_sperm_genes_function_hpo_prot.tsv

Outputs:
- An .xlsx workbook with a gene-centric master sheet and supporting sheets.

All inputs are tab-separated (TSV). All arguments are named.
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
from pathlib import Path
from typing import Iterable, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def parse_args() -> argparse.Namespace:
    """
    Parse command line arguments.

    Returns
    -------
    argparse.Namespace
        Parsed arguments.
    """
    parser = argparse.ArgumentParser(
        description="Build a master fertility evidence workbook from results/ folders."
    )
    parser.add_argument(
        "--base_dir",
        required=True,
        type=Path,
        help="Base project directory, e.g. /home/pthorpe001/data/2026_sperm_Gates",
    )
    parser.add_argument(
        "--testis_run_id",
        required=True,
        type=str,
        help=(
            "Testis pipeline run folder under results/, "
            "e.g. testis_tau0.95_testisTPM5_presentTPM5_spermTPM0.1"
        ),
    )
    parser.add_argument(
        "--out_xlsx",
        required=True,
        type=Path,
        help="Output Excel workbook path (.xlsx).",
    )
    parser.add_argument(
        "--include_variant_sheets",
        action="store_true",
        help="Include variant-level sheets (may increase file size).",
    )
    return parser.parse_args()


def read_tsv(path: Path) -> pd.DataFrame:
    """
    Read a TSV file safely.

    Parameters
    ----------
    path : Path
        TSV path.

    Returns
    -------
    pd.DataFrame
        DataFrame with string dtype where possible.
    """
    return pd.read_csv(path, sep="\t", dtype=str, low_memory=False)


def normalise_gene_symbol(value: str) -> str:
    """
    Normalise a gene symbol for joins.

    Parameters
    ----------
    value : str
        Gene symbol.

    Returns
    -------
    str
        Uppercase symbol with whitespace removed.
    """
    if value is None:
        return ""
    v = str(value).strip()
    v = re.sub(r"\s+", "", v)
    return v.upper()


def ensure_gene_key(df: pd.DataFrame, candidate_cols: Iterable[str]) -> pd.DataFrame:
    """
    Ensure a DataFrame contains 'gene_key' derived from a candidate symbol column.

    Parameters
    ----------
    df : pd.DataFrame
        Input DataFrame.
    candidate_cols : Iterable[str]
        Candidate columns to use for gene symbol.

    Returns
    -------
    pd.DataFrame
        Copy of df with 'gene_key' column.
    """
    for col in candidate_cols:
        if col in df.columns:
            out = df.copy()
            out["gene_key"] = out[col].fillna("").map(normalise_gene_symbol)
            return out
    raise ValueError(
        f"Could not find gene symbol column in {list(candidate_cols)}. "
        f"Available columns: {list(df.columns)}"
    )


def classify_proteomics_evidence(df: pd.DataFrame) -> pd.DataFrame:
    """
    Add a derived proteomics evidence level column based on proteomics support.

    Rules
    -----
    - None: prot_present_any is False/empty.
    - Strong: present and (unique peptides >= 2) and (coverage >= 10).
    - Detected: present but does not meet Strong thresholds.

    Parameters
    ----------
    df : pd.DataFrame
        Gene-level DataFrame containing proteomics columns:
        prot_present_any, prot_unique_peptides_max, prot_coverage_pct_max.

    Returns
    -------
    pd.DataFrame
        Copy of df with proteomics_evidence_level column added.
    """
    out = df.copy()

    def _to_float(value: object) -> float:
        if value is None:
            return 0.0
        s = str(value).strip()
        if s == "" or s.upper() == "NA":
            return 0.0
        try:
            return float(s)
        except ValueError:
            return 0.0

    def _to_bool(value: object) -> bool:
        if value is None:
            return False
        s = str(value).strip().lower()
        return s in {"true", "t", "1", "yes", "y"}

    levels: List[str] = []
    for _, row in out.iterrows():
        present = _to_bool(row.get("prot_present_any", ""))
        uniq = _to_float(row.get("prot_unique_peptides_max", ""))
        cov = _to_float(row.get("prot_coverage_pct_max", ""))

        if not present:
            levels.append("None")
        elif uniq >= 2.0 and cov >= 10.0:
            levels.append("Strong")
        else:
            levels.append("Detected")

    out["proteomics_evidence_level"] = levels
    return out


def summarise_clinvar_by_gene(df: pd.DataFrame) -> pd.DataFrame:
    """
    Summarise ClinVar variant rows to gene-level counts.

    Parameters
    ----------
    df : pd.DataFrame
        ClinVar TSV DataFrame (variant rows).

    Returns
    -------
    pd.DataFrame
        Gene-level summary counts.
    """
    if "GeneSymbol" not in df.columns:
        raise ValueError("ClinVar table must contain 'GeneSymbol' column.")
    work = ensure_gene_key(df, candidate_cols=["GeneSymbol"])

    group = work.groupby("gene_key", dropna=False)

    out = pd.DataFrame(
        {
            "gene_key": group.size().index,
            "n_rows": group.size().values.astype(int),
            "n_unique_allele_id": group["AlleleID"].nunique(dropna=True).values.astype(int)
            if "AlleleID" in work.columns
            else group.size().values.astype(int),
            "n_unique_variation_id": group["VariationID"].nunique(dropna=True).values.astype(int)
            if "VariationID" in work.columns
            else group.size().values.astype(int),
        }
    )

    if "ClinicalSignificance" in work.columns:
        cs = (
            work.groupby(["gene_key", "ClinicalSignificance"])
            .size()
            .reset_index(name="n")
        )
        cs_pivot = cs.pivot_table(
            index="gene_key", columns="ClinicalSignificance", values="n", fill_value=0
        ).reset_index()
        cs_pivot.columns = [
            "gene_key" if c == "gene_key" else f"cs__{c}" for c in cs_pivot.columns
        ]
        out = out.merge(cs_pivot, on="gene_key", how="left")

    if "ReviewStatus" in work.columns:
        rs = (
            work.groupby(["gene_key", "ReviewStatus"])
            .size()
            .reset_index(name="n")
        )
        rs_pivot = rs.pivot_table(
            index="gene_key", columns="ReviewStatus", values="n", fill_value=0
        ).reset_index()
        rs_pivot.columns = [
            "gene_key" if c == "gene_key" else f"rs__{c}" for c in rs_pivot.columns
        ]
        out = out.merge(rs_pivot, on="gene_key", how="left")

    return out


def write_sheet(wb: Workbook, title: str, df: pd.DataFrame) -> None:
    """
    Write a DataFrame to an Excel worksheet with basic formatting.

    Parameters
    ----------
    wb : Workbook
        Openpyxl workbook.
    title : str
        Sheet name.
    df : pd.DataFrame
        Data to write.
    """
    ws = wb.create_sheet(title=title)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(vertical="center", wrap_text=True)

    ws.append(list(df.columns))
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for _, row in df.iterrows():
        ws.append([row.get(c, "") for c in df.columns])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, col_name in enumerate(df.columns, start=1):
        sample = df[col_name].astype(str).head(200).tolist()
        max_len = max([len(str(col_name))] + [len(v) for v in sample])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 60)


def main() -> None:
    """
    Entry point.
    """
    args = parse_args()

    base_dir = args.base_dir
    results_dir = base_dir / "results"
    testis_run_dir = results_dir / args.testis_run_id

    # HPO outputs
    hpo_dir = results_dir / "01_hpo"
    hpo_genes_summary = hpo_dir / "male_infertility_genes_summary.tsv"
    hpo_gene_symbols = hpo_dir / "gene_lists" / "male_infertility_gene_symbols.tsv"

    # ClinVar tiers + reports
    clinvar_best = results_dir / "04_clinvar_collapsed" / "male_infertility_clinvar_variants_best.tsv"
    clinvar_best_pathogenic = (
        results_dir / "04_clinvar_collapsed" / "male_infertility_clinvar_variants_best_pathogenic.tsv"
    )
    clinvar_hc = (
        results_dir
        / "05_clinvar_high_confidence"
        / "male_infertility_clinvar_variants_high_confidence.tsv"
    )
    clinvar_hc_pathogenic = (
        results_dir
        / "05_clinvar_high_confidence"
        / "male_infertility_clinvar_variants_high_confidence_pathogenic.tsv"
    )
    clinvar_hc_gene_summary = (
        results_dir
        / "06_reports"
        / "male_infertility_clinvar_pathogenic_high_confidence_gene_summary.tsv"
    )
    clinvar_hc_report = (
        results_dir
        / "06_reports"
        / "male_infertility_clinvar_pathogenic_high_confidence_report.tsv"
    )

    # Testis integrated table + final high-confidence gene set
    testis_annotated = (
        testis_run_dir
        / "08_proteomics_annotation"
        / "gtex_v11_testis_ranked_with_sperm_presence_function_hpo_prot.tsv"
    )
    testis_high_conf_final = (
        testis_run_dir
        / "09_high_confidence_final"
        / "high_confidence_testis_sperm_genes_function_hpo_prot.tsv"
    )
    tissue_medians = testis_run_dir / "01_gtex_specificity" / "gtex_v11_tissue_medians.tsv"

    required_paths: List[Path] = [
        hpo_genes_summary,
        hpo_gene_symbols,
        clinvar_best,
        clinvar_best_pathogenic,
        clinvar_hc,
        clinvar_hc_pathogenic,
        clinvar_hc_gene_summary,
        clinvar_hc_report,
        testis_annotated,
        testis_high_conf_final,
    ]
    missing = [p for p in required_paths if not p.exists() or p.stat().st_size == 0]
    if missing:
        raise FileNotFoundError("Missing required inputs:\n" + "\n".join(str(p) for p in missing))

    # Load
    testis_df = read_tsv(testis_annotated)
    testis_df = ensure_gene_key(
        testis_df,
        candidate_cols=[
            "gene_symbol_norm",
            "hgnc_symbol_norm",
            "hgnc_symbol",
            "hgnc_symbol_from_ensembl",
        ],
    )

    hpo_summary_df = read_tsv(hpo_genes_summary)
    hpo_summary_df = ensure_gene_key(hpo_summary_df, candidate_cols=["gene_symbol", "GeneSymbol", "symbol"])

    hpo_symbols_df = read_tsv(hpo_gene_symbols)
    hpo_symbols_df = ensure_gene_key(hpo_symbols_df, candidate_cols=["gene_symbol", "GeneSymbol", "symbol"])

    clinvar_best_df = read_tsv(clinvar_best)
    clinvar_best_df = ensure_gene_key(clinvar_best_df, candidate_cols=["GeneSymbol"])

    clinvar_best_path_df = read_tsv(clinvar_best_pathogenic)
    clinvar_best_path_df = ensure_gene_key(clinvar_best_path_df, candidate_cols=["GeneSymbol"])

    clinvar_hc_df = read_tsv(clinvar_hc)
    clinvar_hc_df = ensure_gene_key(clinvar_hc_df, candidate_cols=["GeneSymbol"])

    clinvar_hc_path_df = read_tsv(clinvar_hc_pathogenic)
    clinvar_hc_path_df = ensure_gene_key(clinvar_hc_path_df, candidate_cols=["GeneSymbol"])

    clinvar_hc_gene_summary_df = read_tsv(clinvar_hc_gene_summary)
    clinvar_hc_gene_summary_df = ensure_gene_key(
        clinvar_hc_gene_summary_df, candidate_cols=["GeneSymbol", "Gene", "gene_symbol"]
    )

    clinvar_hc_report_df = read_tsv(clinvar_hc_report)

    testis_hc_final_df = read_tsv(testis_high_conf_final)
    testis_hc_final_df = ensure_gene_key(
        testis_hc_final_df,
        candidate_cols=["gene_symbol_norm", "hgnc_symbol_norm", "hgnc_symbol"],
    )

    # Summarise ClinVar per tier (gene-level)
    best_gene = summarise_clinvar_by_gene(clinvar_best_df).rename(
        columns={c: f"clinvar_best__{c}" for c in summarise_clinvar_by_gene(clinvar_best_df).columns if c != "gene_key"}
    )
    best_path_gene = summarise_clinvar_by_gene(clinvar_best_path_df).rename(
        columns={c: f"clinvar_best_pathogenic__{c}" for c in summarise_clinvar_by_gene(clinvar_best_path_df).columns if c != "gene_key"}
    )
    hc_gene = summarise_clinvar_by_gene(clinvar_hc_df).rename(
        columns={c: f"clinvar_hc__{c}" for c in summarise_clinvar_by_gene(clinvar_hc_df).columns if c != "gene_key"}
    )
    hc_path_gene = summarise_clinvar_by_gene(clinvar_hc_path_df).rename(
        columns={c: f"clinvar_hc_pathogenic__{c}" for c in summarise_clinvar_by_gene(clinvar_hc_path_df).columns if c != "gene_key"}
    )

    # Genes master: start from the integrated testis table
    genes_master = testis_df.copy()

    # Tier flags
    genes_master["in_hpo_gene_set"] = genes_master["gene_key"].isin(set(hpo_symbols_df["gene_key"]))
    genes_master["clinvar_best_present"] = genes_master["gene_key"].isin(set(clinvar_best_df["gene_key"]))
    genes_master["clinvar_best_pathogenic_present"] = genes_master["gene_key"].isin(set(clinvar_best_path_df["gene_key"]))
    genes_master["clinvar_hc_present"] = genes_master["gene_key"].isin(set(clinvar_hc_df["gene_key"]))
    genes_master["clinvar_hc_pathogenic_present"] = genes_master["gene_key"].isin(set(clinvar_hc_path_df["gene_key"]))
    genes_master["in_testis_high_conf_final"] = genes_master["gene_key"].isin(set(testis_hc_final_df["gene_key"]))

    # Derived summary column for presentation-friendly proteomics interpretation

    genes_master = classify_proteomics_evidence(df=genes_master)



    # Merge HPO gene summary fields (prefixed)
    hpo_cols = [c for c in hpo_summary_df.columns if c not in ["gene_key"]]
    hpo_ren = {c: f"hpo__{c}" for c in hpo_cols}
    genes_master = genes_master.merge(hpo_summary_df.rename(columns=hpo_ren), on="gene_key", how="left")

    # Merge ClinVar gene summaries
    genes_master = genes_master.merge(best_gene, on="gene_key", how="left")
    genes_master = genes_master.merge(best_path_gene, on="gene_key", how="left")
    genes_master = genes_master.merge(hc_gene, on="gene_key", how="left")
    genes_master = genes_master.merge(hc_path_gene, on="gene_key", how="left")

    # Add ClinVar HC pathogenic gene summary (the 24/25-gene list output)
    hc_gs_cols = [c for c in clinvar_hc_gene_summary_df.columns if c != "gene_key"]
    hc_gs_ren = {c: f"clinvar_hc_pathogenic_gene_summary__{c}" for c in hc_gs_cols}
    genes_master = genes_master.merge(
        clinvar_hc_gene_summary_df.rename(columns=hc_gs_ren),
        on="gene_key",
        how="left",
    )

    # Build a compact tier summary sheet (one row per gene, with the tier flags)
    tier_summary_cols = [
        "gene_key",
        "in_hpo_gene_set",
        "clinvar_best_present",
        "clinvar_best_pathogenic_present",
        "clinvar_hc_present",
        "clinvar_hc_pathogenic_present",
        "in_testis_high_conf_final",
    ]
    tier_summary = genes_master[tier_summary_cols].copy()

    prot_anchor = "prot_present_fraction"
    if "proteomics_evidence_level" in genes_master.columns and prot_anchor in genes_master.columns:
        cols = list(genes_master.columns)
        cols.remove("proteomics_evidence_level")
        insert_at = cols.index(prot_anchor) + 1
        cols.insert(insert_at, "proteomics_evidence_level")
        genes_master = genes_master[cols]


    # Workbook
    wb = Workbook()
    wb.remove(wb.active)

    readme = pd.DataFrame(
        {
            "field": [
                "created",
                "base_dir",
                "results_dir",
                "testis_run_id",
                "testis_annotated_table",
                "testis_high_conf_final",
                "hpo_genes_summary",
                "hpo_gene_symbols",
                "clinvar_best",
                "clinvar_best_pathogenic",
                "clinvar_high_confidence",
                "clinvar_high_confidence_pathogenic",
                "clinvar_hc_pathogenic_gene_summary",
                "clinvar_hc_pathogenic_report",
                "include_variant_sheets",
            ],
            "value": [
                dt.datetime.now().isoformat(timespec="seconds"),
                str(base_dir),
                str(results_dir),
                args.testis_run_id,
                str(testis_annotated),
                str(testis_high_conf_final),
                str(hpo_genes_summary),
                str(hpo_gene_symbols),
                str(clinvar_best),
                str(clinvar_best_pathogenic),
                str(clinvar_hc),
                str(clinvar_hc_pathogenic),
                str(clinvar_hc_gene_summary),
                str(clinvar_hc_report),
                str(bool(args.include_variant_sheets)),
            ],
        }
    )
    write_sheet(wb, "README", readme)
    write_sheet(wb, "Genes_Master", genes_master)
    write_sheet(wb, "ClinVar_Tier_Summary", tier_summary)
    write_sheet(wb, "HPO_Genes_Summary", hpo_summary_df)
    write_sheet(wb, "Testis_HighConfidence_Final", testis_hc_final_df)

    if tissue_medians.exists() and tissue_medians.stat().st_size > 0:
        tissue_df = read_tsv(tissue_medians)
        write_sheet(wb, "GTEx_Tissue_Medians", tissue_df)

    # Variant-level sheets (optional)
    if args.include_variant_sheets:
        write_sheet(wb, "ClinVar_Best_Variants", clinvar_best_df)
        write_sheet(wb, "ClinVar_Best_Pathogenic_Variants", clinvar_best_path_df)
        write_sheet(wb, "ClinVar_HC_Variants", clinvar_hc_df)
        write_sheet(wb, "ClinVar_HC_Pathogenic_Variants", clinvar_hc_path_df)
        write_sheet(wb, "ClinVar_HC_Pathogenic_Report", clinvar_hc_report_df)

    args.out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out_xlsx)


if __name__ == "__main__":
    main()
